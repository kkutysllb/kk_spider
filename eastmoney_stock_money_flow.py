#!/usr/bin/env python
# _*_ encoding: utf-8 _*_
# @Author kkutysllb
import os.path
import random
import time

import openpyxl
import pymysql
import requests
import json
import jsonpath
import pandas as pd
import datetime
import urllib3
from openpyxl import load_workbook, Workbook
from sqlalchemy import create_engine

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def create_params(page):
    params = {
        'pz': '50',
        'po': '1',
        'np': '1',
        'ut': 'b2884a393a59ad64002292a3e90d46a5',
        'fltt': '2',
        'invt': '2',
        'fid': 'f62',
        'fs': 'm:0+t:6+f:!2,m:0+t:13+f:!2,m:0+t:80+f:!2,m:1+t:2+f:!2,m:1+t:23+f:!2,m:0+t:7+f:!2,m:1+t:3+f:!2',
        'fields': 'f12,f14,f2,f3,f62,f184,f66,f69,f72,f75,f78,f81,f84,f87,f204,f205,f124,f1,f13',
        # 'cb': 'jQuery112307249951414423113_1647790279433',
    }
    value = str(page)
    params['pn'] = value
    # print(params)
    return params


def get_response(params):
    cookies = {
        'st_si': '99052079014232',
        'st_asi': 'delete',
        'cowCookie': 'true',
        'qgqp_b_id': '1534648bc0459d8e8626fd271b926c60',
        'intellpositionL': '249px',
        'intellpositionT': '755px',
        'st_pvi': '88918293181473',
        'st_sp': '2022-03-02%2018%3A20%3A49',
        'st_inirUrl': 'https%3A%2F%2Flink.csdn.net%2F',
        'st_sn': '44',
        'st_psi': '20220320233050460-113300300813-6213078942',
        'em_hq_fls': 'js',
        'HAList': 'a-sz-301159-%u4E09%u7EF4%u5929%u5730%2Ca-sz-300059-%u4E1C%u65B9%u8D22%u5BCC%2Cty-0-399001-%u6DF1%u8BC1%u6210%u6307',
    }

    headers = {
        'Connection': 'keep-alive',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.74 Safari/537.36',
        'Accept': '*/*',
        'Referer': 'http://data.eastmoney.com/',
        'Accept-Language': 'zh-CN,zh-TW;q=0.9,zh;q=0.8,en-US;q=0.7,en;q=0.6',
    }
    response = requests.get('https://push2.eastmoney.com/api/qt/clist/get', headers=headers, params=params,
                            cookies=cookies, verify=False)
    return response


def get_data(res):
    content = res.text
    data = json.loads(content)
    # data = json.dumps(data)
    return data


def data_cleaning(data):
    # ????????????
    stock_code_list = jsonpath.jsonpath(data, '$..f12')
    new_stock_code_list = []
    for code in stock_code_list:
        new_stock_code_list.append('ts'+code)
    # ????????????
    name_list = jsonpath.jsonpath(data, '$..f14')
    # ?????????
    latest_price_list = jsonpath.jsonpath(data, '$..f2')
    # ?????????
    price_limit_list = jsonpath.jsonpath(data, '$..f3')
    # ???????????????
    net_amount_list1 = jsonpath.jsonpath(data, '$..f62')
    # net_proportion_list1 = jsonpath.jsonpath(data, '$..f184')
    # print(net_proportion_list1)
    # exit()
    # ??????????????????
    net_amount_list2 = jsonpath.jsonpath(data, '$..f66')
    # net_proportion_list2 = jsonpath.jsonpath(data, '$..f69')
    # ???????????????
    net_amount_list3 = jsonpath.jsonpath(data, '$..f72')
    # net_proportion_list3 = jsonpath.jsonpath(data, '$..f75')
    # ???????????????
    net_amount_list4 = jsonpath.jsonpath(data, '$..f78')
    # net_proportion_list4 = jsonpath.jsonpath(data, '$..f81')
    # ???????????????
    net_amount_list5 = jsonpath.jsonpath(data, '$..f84')
    # net_proportion_list5 = jsonpath.jsonpath(data, '$..f87')
    date_list = []
    for item in range(len(stock_code_list)):
        date_list.append(str(datetime.datetime.now()).split(' ')[0])
    # print(date_list, len(date_list))
    df = pd.DataFrame(new_stock_code_list, columns=['code'])
    df['date'] = date_list
    # df['code'] = stock_code_list
    df['name'] = name_list
    df['new_price'] = latest_price_list
    df['price_pct'] = price_limit_list
    df['net_amount'] = net_amount_list1
    # df['???????????????-?????????(%)'] = net_proportion_list1
    df['net_super_amount'] = net_amount_list2
    # df['??????????????????-?????????(%)'] = net_proportion_list2
    df['net_large_amount'] = net_amount_list3
    # df['???????????????-?????????(%)'] = net_proportion_list3
    df['net_middle_amount'] = net_amount_list4
    # df['???????????????-?????????(%)'] = net_proportion_list4
    df['net_small_amount'] = net_amount_list5
    # df['???????????????-?????????(%)'] = net_proportion_list5

    df.set_index(['date'], inplace=True)
    return df


def save_csv(data):
    file_root = './stock_money_datas/'
    filename = file_root + 'eastmoney_stocks_' + str(datetime.datetime.now()).split(' ')[0] + '.csv'

    if os.path.exists(filename):
        data.to_csv(filename, mode='a', header=False, encoding='utf-8')
    else:
        data.to_csv(filename, encoding='utf-8')
    print('????????????!')


def save_excel(data, page):
    file_root = './stock_money_datas/'
    filename = file_root + 'eastmoney_stocks_' + str(datetime.datetime.now()).split(' ')[0] + '.xlsx'
    if os.path.exists(filename):
        wb = load_workbook(filename)
        # wb.save(filename)
    else:
        wb = Workbook(filename)
        wb.save(filename)
    # book = load_workbook('east_money.xlsx')
    with pd.ExcelWriter(filename) as E:
        E.book = wb
        E.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        sheet_name = '???' + str(page) + '???'
        data.to_excel(E, sheet_name=sheet_name)


def from_files_to_save_db(filename, table_name, ):
    db = pymysql.connect(host='localhost',
                         user='root',
                         password='Oms_2600',
                         db='kk_trader_db',
                         port=3306)
    cursor = db.cursor()

    df = pd.read_csv(filename, encoding='utf-8')
    # ??????????????????
    keys = df.keys()
    values = df.values.tolist()

    key_sql = ','.join(keys)
    value_sql = ','.join(['%s'] * df.shape[1])

    # ????????????
    insert_data_str = """ insert into %s (%s) values (%s)""" % (table_name, key_sql, value_sql)

    # ?????????????????????
    cursor.executemany(insert_data_str, values)
    db.commit()

    # ????????????
    cursor.close()
    db.close()


if __name__ == '__main__':
    # for page in range(1, 96):
    #     params = create_params(page)
    #     response = get_response(params)
    #     data = get_data(response)
    #     df = data_cleaning(data)
    #     save_csv(df)

    time.sleep(random.randint(1, 2))
    from_files_to_save_db('./stock_money_datas/eastmoney_stocks_2022-03-21.csv', 't_eastmoney_stocks_money_flow')
